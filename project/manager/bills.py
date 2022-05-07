from datetime import datetime
import os

from django.http import HttpResponseRedirect, HttpResponse
from django.views import View
from django.db.models import Q, Sum, Prefetch, F, Avg, Case, When, Value, FloatField, IntegerField
from django.db.models.functions import Round

from project.manager import models

from openpyxl import Workbook, drawing
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

class CreateExcel(View):

    def font_h1(self):
        return Font(name='Times New Roman',
                    size=14,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

    def font_h2(self):
        return Font(name='Times New Roman',
                    size=12,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

    def font_p1(self):
        return Font(name='Times New Roman',
                    size=14,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
    def font_p2(self):
        return Font(name='Times New Roman',
                    size=12,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

    def font_foot(self):
        return Font(name='Times New Roman',
                    size=12,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF363742')

    def init_sheet(self, sheet, number):
        company = models.Company.objects.get(id=1)
        # h1
        sheet['D24'] = "Total TTC"
        sheet['D24'].font = self.font_h1()
        sheet['E32'] = company.cellphone_comp
        sheet['E32'].font = self.font_h1()
        sheet['E32'].alignment = Alignment(horizontal="center")
        # h2
        sheet['D12'] = "Chambéry, le"
        sheet['D12'].font = self.font_h2()
        sheet['C14'] = number
        sheet['C14'].font = self.font_h2()
        sheet['C16'] = "Nombre de repas"
        sheet['C16'].font = self.font_h2()
        sheet['C18'] = "Prix unitaire ttc"
        sheet['C18'].font = self.font_h2()
        # p1
        sheet['B14'] = "Facture N°"
        sheet['B14'].font = self.font_p1()
        sheet['D21'] = "Total HT"
        sheet['D21'].font = self.font_p1()
        sheet['D22'] = "TVA 5.5%"
        sheet['D22'].font = self.font_p1()
        # p2
        text = company.comment_comp
        number_of_lines = len(text) // 80 + 1
        initial_row = 28
        put_to_other_line = ''
        for row_offset in range(number_of_lines):
            text_line = put_to_other_line + text[row_offset * 80:(row_offset + 1) * 80]
            put_to_other_line = ''
            if text_line[-1] != ' ' and row_offset != number_of_lines - 1:
                put_to_other_line = text_line.split(' ')[-1]
                text_line = ' '.join(text_line.split(' ')[:-1])

            sheet[f'E{initial_row + row_offset}'] = text_line
            sheet[f'E{initial_row + row_offset}'].font = self.font_p2()
            sheet[f'E{initial_row + row_offset}'].alignment = Alignment(horizontal="center")
        # foot
        sheet['E34'] = company.company_comp
        sheet['E34'].font = self.font_foot()
        sheet['E34'].alignment = Alignment(horizontal="center")
        sheet['E35'] = company.address_comp
        sheet['E35'].font = self.font_foot()
        sheet['E35'].alignment = Alignment(horizontal="center")
        sheet['E36'] = company.siret_comp
        sheet['E36'].font = self.font_foot()
        sheet['E36'].alignment = Alignment(horizontal="center")

        return sheet

    def post(self, request, *args, **kwargs):
        # init
        wb = Workbook()
        month = self.request.POST.get('month', datetime.now().month)
        year = self.request.POST.get('year', datetime.now().year)
        date_posted = self.request.POST.get('date', "")
        if date_posted != "":
            dates = date_posted.split('-')
            date = '-'.join([dates[-1], dates[1], dates[0]])
        else:
            date = f"{datetime.now().year}-{datetime.now().month}-{datetime.now().day}"
        clients = models.Client.objects.all().order_by('circuit').order_by('last_name', 'first_name')
        commands = models.Command.objects.filter(year_date_command=year, month_date_command=month)

        for client in clients:
            # Don't display clients that doesn't pay this month
            customer_this_month = commands.filter(client=client)\
                    .aggregate(sum=Sum(
                        (F('meals__default__price')) * F('command_command'),
                        output_field=FloatField(),
                        )
                    )['sum']
            if not customer_this_month:
                continue

            # Elephant img added
            img = drawing.image.Image(os.path.join(os.getcwd(), 'project', 'static', 'img_elephant.png'))
            img.anchor = 'C3'
            img.height = 148
            img.width  = 115

            # Create sheet
            wb.create_sheet(f"{client.id} - {client.last_name} {client.first_name}")
            active_sheet = wb[f"{client.id} - {client.last_name} {client.first_name}"]
            active_sheet = self.init_sheet(active_sheet, month)

            # Insert image
            active_sheet.add_image(img)

            # Insert date
            active_sheet['F12'] = date
            active_sheet['F12'].font = self.font_h2()

            # set clients address
            active_sheet['H6'] = f"{client.last_name} {client.first_name}"
            active_sheet['H6'].font = self.font_h1()
            active_sheet['H7'] = f"{client.address}"
            active_sheet['H7'].alignment = Alignment(wrap_text=True, shrink_to_fit=False)
            active_sheet['H8'] = f"{client.postcode}"

            # Number of meals
            number_of_meals = commands.filter(client=client, command_command__gt=0).aggregate(sum=Sum(F('command_command')))['sum']
            active_sheet['F16'] = number_of_meals if number_of_meals else 0
            active_sheet['F16'].font = self.font_h2()

            
            unit_price = models.Food.objects.filter(id__in=client.client_command.all()).aggregate(sum=(Sum(F('price'))))['sum']
            active_sheet['F18'] = unit_price if unit_price else 0
            active_sheet['F18'].font = self.font_h2()
            active_sheet['F18'].number_format = '#,##0.00€' 

            # TTC calc
            TTC = commands.filter(client=client)\
                    .aggregate(sum=Sum(
                        (F('meals__default__price')) * F('command_command'),
                        output_field=FloatField(),
                        )
                    )['sum']
            if not TTC:
                TTC = 0.0
            active_sheet['F24'] = TTC
            active_sheet['F24'].font = self.font_h2()
            active_sheet['F21'] = "=F24*(1-0.055)"
            active_sheet['F21'].font = self.font_h2()
            active_sheet['F22'] = "=F24*0.055"
            active_sheet['F22'].font = self.font_h2()

            # format €
            active_sheet['F21'].number_format = '#,##0.00€' 
            active_sheet['F22'].number_format = '#,##0.00€' 
            active_sheet['F24'].number_format = '#,##0.00€' 

            # dimensions
            dims = {
                'A': 5,
                'B': 13,
                'C': 4,
                'D': 10,
                'E': 5,
                'F': 10,
                'G': 4,
                'H': 25,
            }
            for col, value in dims.items():
                active_sheet.column_dimensions[col].width = value
            # Border on prices
            thin = Side(border_style="double")
            for number in range(21, 25):
                active_sheet[f'C{number}'].border = Border(left=thin)
                active_sheet[f'G{number}'].border = Border(right=thin)
            for letter in ['D', 'E', 'F']:
                active_sheet[f'{letter}20'].border = Border(top=thin)
                active_sheet[f'{letter}25'].border = Border(bottom=thin)
            active_sheet[f'C20'].border = Border(top=thin, left=thin)
            active_sheet[f'G20'].border = Border(top=thin, right=thin)
            active_sheet[f'C25'].border = Border(bottom=thin, left=thin)
            active_sheet[f'G25'].border = Border(bottom=thin, right=thin)
        # Delete std Sheet
        std = wb.get_sheet_by_name('Sheet')
        wb.remove_sheet(std)

        # Export
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Factures_{month}_{year}.xlsx"'

        return response


class CreateUnitExcel(View):

    def font_h1(self):
        return Font(name='Times New Roman',
                    size=14,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

    def font_h2(self):
        return Font(name='Times New Roman',
                    size=12,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

    def font_p1(self):
        return Font(name='Times New Roman',
                    size=14,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
    def font_p2(self):
        return Font(name='Times New Roman',
                    size=12,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

    def font_foot(self):
        return Font(name='Times New Roman',
                    size=12,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF363742')

    def init_sheet(self, sheet, number):
        company = models.Company.objects.get(id=1)
        # h1

        # h2
        sheet['D12'] = "Chambéry, le"
        sheet['D12'].font = self.font_h2()
        sheet['E14'] = number
        sheet['E14'].font = self.font_h2()
        sheet['C16'] = "Nombre de repas"
        sheet['C16'].font = self.font_h2()
        # p1
        sheet['B14'] = " Détails de la facture N°"
        sheet['B14'].font = self.font_p1()


        return sheet

    def post(self, request, *args, **kwargs):
        # init
        wb = Workbook()
        month = self.request.POST.get('month_unit', datetime.now().month)
        year = self.request.POST.get('year_unit', datetime.now().year)
        client_id = self.request.POST.get('client_unit_bill')
        date_posted = self.request.POST.get('date_unit', "")
        if date_posted != "":
            dates = date_posted.split('-')
            date = '-'.join([dates[-1], dates[1], dates[0]])
        else:
            date = f"{datetime.now().year}-{datetime.now().month}-{datetime.now().day}"
        clients = models.Client.objects.filter(id=client_id)
        commands = models.Command.objects.filter(year_date_command=year, month_date_command=month)

        for client in clients:
            # Don't display clients that doesn't pay this month
            commands = commands.filter(client=client)
            customer_this_month = commands.aggregate(sum=Sum(
                        (F('meals__default__price')) * F('command_command'),
                        output_field=FloatField(),
                        )
                    )['sum']
            if not customer_this_month:
                wb.create_sheet(f"{client.id} - {client.last_name} {client.first_name}")
                continue

            # Elephant img added
            img = drawing.image.Image(os.path.join(os.getcwd(), 'project', 'static', 'img_elephant.png'))
            img.anchor = 'B3'
            img.height = 148
            img.width  = 115

            # Create sheet
            wb.create_sheet(f"{client.id} - {client.last_name} {client.first_name}")
            active_sheet = wb[f"{client.id} - {client.last_name} {client.first_name}"]
            active_sheet = self.init_sheet(active_sheet, month)

            # Insert image
            active_sheet.add_image(img)

            # Insert date
            active_sheet['F12'] = date
            active_sheet['F12'].font = self.font_h2()

            # set clients address
            active_sheet['F6'] = f"{client.last_name} {client.first_name}"
            active_sheet['F6'].font = self.font_h1()
            active_sheet['F7'] = f"{client.address}"
            active_sheet['F7'].alignment = Alignment(wrap_text=True, shrink_to_fit=False)
            active_sheet['F8'] = f"{client.postcode}"

            active_sheet['C16'] = "Plat"
            active_sheet['D16'] = "Prix /u TTC"
            active_sheet['E16'] = "Quantité"
            active_sheet['F16'] = "Prix total TTC"
            active_sheet['C16'].font = self.font_p2()
            active_sheet['C16'].alignment = Alignment(horizontal="center")
            active_sheet['D16'].font = self.font_p2()
            active_sheet['D16'].alignment = Alignment(horizontal="center")
            active_sheet['E16'].font = self.font_p2()
            active_sheet['E16'].alignment = Alignment(horizontal="center")
            active_sheet['F16'].font = self.font_p2()
            active_sheet['F16'].alignment = Alignment(horizontal="center")

            offsetY = 0
            # details
            for food in models.Food.objects.all():
                price = float(str(food.price).replace(',', '.'))
                command_command =commands\
                                    .filter(meals__id=food.id)\
                                    .aggregate(sum=Round(
                                                       Sum(
                                                           F('command_command'),
                                                               output_field=IntegerField(),
                                                       ),
                                                   2
                                                )
                                    )['sum']
                if not command_command:
                    continue
                active_sheet[f'C{17+offsetY}'] = food.category
                active_sheet[f'D{17+offsetY}'] = price
                active_sheet[f'D{17+offsetY}'].number_format = '#,##0.00€' 
                active_sheet[f'D{17+offsetY}'].alignment = Alignment(horizontal="center")
                active_sheet[f'E{17+offsetY}'] = command_command               
                active_sheet[f'E{17+offsetY}'].alignment = Alignment(horizontal="center")              
                active_sheet[f'F{17+offsetY}'] = f"=D{17+offsetY}*E{17+offsetY}"
                active_sheet[f'F{17+offsetY}'].number_format = '#,##0.00€' 
                active_sheet[f'F{17+offsetY}'].alignment = Alignment(horizontal="center")
                offsetY += 1
            #Totals
            active_sheet[f'D{17 + offsetY + 1}'] = "Total"
            active_sheet[f'E{17 + offsetY + 1}'] = f"=SUM(E{17}:E{17+offsetY})"
            active_sheet[f'F{17 + offsetY + 1}'] = f"=SUM(F{17}:F{17+offsetY})"
            active_sheet[f'D{17 + offsetY + 1}'].font = self.font_h1()
            active_sheet[f'D{17 + offsetY + 1}'].alignment = Alignment(horizontal="center")
            active_sheet[f'E{17 + offsetY + 1}'].font = self.font_h1()
            active_sheet[f'E{17 + offsetY + 1}'].alignment = Alignment(horizontal="center")
            active_sheet[f'F{17 + offsetY + 1}'].font = self.font_h1()
            active_sheet[f'F{17 + offsetY + 1}'].number_format = '#,##0.00€' 
            active_sheet[f'F{17 + offsetY + 1}'].alignment = Alignment(horizontal="center")

            
            company = models.Company.objects.get(id=1)

            # cellphone
            active_sheet[f'E{24 + offsetY + 2}'] = company.cellphone_comp
            active_sheet[f'E{24 + offsetY + 2}'].font = self.font_h1()
            active_sheet[f'E{24 + offsetY + 2}'].alignment = Alignment(horizontal="center")

            # footer
            active_sheet[f'E{26 + offsetY + 2}'] = company.company_comp
            active_sheet[f'E{26 + offsetY + 2}'].font = self.font_foot()
            active_sheet[f'E{26 + offsetY + 2}'].alignment = Alignment(horizontal="center")
            active_sheet[f'E{27 + offsetY + 2}'] = company.address_comp
            active_sheet[f'E{27 + offsetY + 2}'].font = self.font_foot()
            active_sheet[f'E{27 + offsetY + 2}'].alignment = Alignment(horizontal="center")
            active_sheet[f'E{28 + offsetY + 2}'] = company.siret_comp
            active_sheet[f'E{28 + offsetY + 2}'].font = self.font_foot()
            active_sheet[f'E{28 + offsetY + 2}'].alignment = Alignment(horizontal="center")

            # p2
            text = company.comment_comp
            number_of_lines = len(text) // 80 + 1
            initial_row = 20 + offsetY
            put_to_other_line = ''
            for row_offset in range(number_of_lines):
                text_line = put_to_other_line + text[row_offset * 80:(row_offset + 1) * 80]
                put_to_other_line = ''
                if text_line[-1] != ' ' and row_offset != number_of_lines - 1:
                    put_to_other_line = text_line.split(' ')[-1]
                    text_line = ' '.join(text_line.split(' ')[:-1])

                active_sheet[f'E{initial_row + row_offset}'] = text_line
                active_sheet[f'E{initial_row + row_offset}'].font = self.font_p2()
                active_sheet[f'E{initial_row + row_offset}'].alignment = Alignment(horizontal="center")
            # dimensions
            dims = {
                'A': 5,
                'B': 5,
                'C': 12,
                'D': 10,
                'E': 9,
                'F': 35,
                'G': 35,
                'H': 35,
            }
            for col, value in dims.items():
                active_sheet.column_dimensions[col].width = value
            # Border on prices
            # thin = Side(border_style="double")
            # for number in range(21, 25):
            #     active_sheet[f'C{number}'].border = Border(left=thin)
            #     active_sheet[f'G{number}'].border = Border(right=thin)
            # for letter in ['D', 'E', 'F']:
            #     active_sheet[f'{letter}20'].border = Border(top=thin)
            #     active_sheet[f'{letter}25'].border = Border(bottom=thin)
            # active_sheet[f'C20'].border = Border(top=thin, left=thin)
            # active_sheet[f'G20'].border = Border(top=thin, right=thin)
            # active_sheet[f'C25'].border = Border(bottom=thin, left=thin)
            # active_sheet[f'G25'].border = Border(bottom=thin, right=thin)
        # Delete std Sheet
        std = wb.get_sheet_by_name('Sheet')
        wb.remove_sheet(std)

        # Export
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Factures_{month}_{year}.xlsx"'

        return response
